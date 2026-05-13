"""
APP.PY — Servidor integrado: YOLO + Web UI + Arduino
======================================================
Captura video, ejecuta YOLOv8, transmite MJPEG con bounding boxes,
envía estadísticas por WebSocket, y controla Arduino por serial.

Uso:
    python APP.PY

Dependencias:
    pip install flask flask-socketio ultralytics opencv-python pyserial
"""

from flask import Flask, render_template, Response, request, jsonify, redirect, url_for, session, flash, send_file
from flask_socketio import SocketIO, emit
from functools import wraps
from ultralytics import YOLO
from collections import deque
from datetime import datetime
import serial
import threading
import hashlib
import numpy as np
import time
import cv2
import os
import io

app = Flask(__name__)
app.config['SECRET_KEY'] = 'secreto_muy_seguro_2024'
socketio = SocketIO(app, cors_allowed_origins='*')

# ─── Usuarios ─────────────────────────────────────────────────────────────────
USUARIOS = {
    'admin':      hashlib.sha256('Adm!n#9x2K'.encode()).hexdigest(),
    'eduardo':    hashlib.sha256('Ed$7mQpR!3'.encode()).hexdigest(),
    'diego':      hashlib.sha256('Di#4kWnZ@8'.encode()).hexdigest(),
    'cesar_gael': hashlib.sha256('CG!6vTxL#2'.encode()).hexdigest(),
    'saul':       hashlib.sha256('Sa@9jBqM!5'.encode()).hexdigest(),
}

# ─── Configuración ────────────────────────────────────────────────────────────
import platform
import sys

ES_RASPBERRY = platform.machine().startswith('aarch64') or platform.machine().startswith('arm')

if ES_RASPBERRY:
    MODELO_PATH   = 'modelo/best_ncnn_model'       # NCNN para ARM
    PUERTO_SERIAL = '/dev/ttyUSB0'                  # Puerto serial en Linux
    print('🍓 Detectado: Raspberry Pi')
else:
    MODELO_PATH   = 'runs/detect/runs/entrenar/tuercas_tornillos/weights/best.pt'
    PUERTO_SERIAL = 'COM7'                          # Puerto serial en Windows
    print('💻 Detectado: Laptop/PC')

CAMARA_INDEX   = 0
RESOLUCION     = (640, 480)
BAUDRATE       = 9600

# ─── Nombres y colores de clases ──────────────────────────────────────────────
NOMBRES_CLASE = {0: 'Tornillo', 1: 'Tuerca', -1: 'Obj. Extrano'}
COLORES_BGR   = {0: (255, 180, 0), 1: (0, 255, 120), -1: (0, 0, 255)}

# ─── Estado global ────────────────────────────────────────────────────────────
estado = {
    'activo': False,
    'fps': 0,
    'config': {
        'velocidad_banda': 50,
        'umbral_confianza': 0.30,
        'umbral_aceptacion': 0.65,
        'fps_objetivo': 15,
    },
    'contadores': {
        'inspeccionadas': 0,
        'aceptadas': 0,
        'rechazadas': 0,
        'objetos_extranos': 0,
    },
}

eventos = deque(maxlen=200)
alarmas = []
alarma_id_counter = 0
metricas_horarias = {}
detecciones_log = []      # Registro detallado para exportar a Excel
sesion_inicio = datetime.now()

frame_actual = None
frame_lock = threading.Lock()
modelo = None
arduino = None

# ─── Tracking para detección por flanco de subida ─────────────────────────────
# Cada objeto trackeado tiene: id, bbox, cls_id, conf, frames_sin_ver
objetos_previos = []       # Lista de objetos activos del frame anterior
siguiente_obj_id = 0       # ID incremental para objetos nuevos
MAX_FRAMES_AUSENTE = 10    # Frames sin ver antes de eliminar un objeto del tracking
IOU_UMBRAL = 0.3           # IoU mínimo para considerar que es el mismo objeto

def calcular_iou(boxA, boxB):
    """Calcula Intersection over Union entre dos bboxes [x1,y1,x2,y2]."""
    xA = max(boxA[0], boxB[0])
    yA = max(boxA[1], boxB[1])
    xB = min(boxA[2], boxB[2])
    yB = min(boxA[3], boxB[3])
    inter = max(0, xB - xA) * max(0, yB - yA)
    if inter == 0:
        return 0.0
    areaA = (boxA[2] - boxA[0]) * (boxA[3] - boxA[1])
    areaB = (boxB[2] - boxB[0]) * (boxB[3] - boxB[1])
    return inter / (areaA + areaB - inter)
motorActivo = False

# ─── Cargar modelo YOLO ──────────────────────────────────────────────────────
def cargar_modelo():
    global modelo
    if not os.path.exists(MODELO_PATH):
        print(f'❌ Modelo no encontrado: {MODELO_PATH}')
        return
    print(f'🔄 Cargando modelo: {MODELO_PATH}')
    modelo = YOLO(MODELO_PATH)
    dummy = np.zeros((640, 640, 3), dtype=np.uint8)
    modelo.predict(dummy, verbose=False)
    print('✅ Modelo YOLOv8 cargado y listo')

# ─── Arduino ──────────────────────────────────────────────────────────────────
def conectar_arduino():
    global arduino
    try:
        arduino = serial.Serial(PUERTO_SERIAL, BAUDRATE, timeout=1)
        time.sleep(2)
        print(f'✅ Arduino conectado en {PUERTO_SERIAL}')
    except Exception as e:
        print(f'⚠️  Arduino no disponible: {e}')
        arduino = None

def enviar_comando(cmd):
    if arduino and arduino.is_open:
        try:
            arduino.write((cmd + '\n').encode())
            print(f'📤 Enviado: {cmd}')
            return True
        except Exception as e:
            print(f'❌ Error serial: {e}')
            return False
    print(f'⚠️  Arduino no conectado — {cmd}')
    return False

def leer_serial():
    while True:
        if arduino and arduino.is_open:
            try:
                if arduino.in_waiting:
                    linea = arduino.readline().decode('utf-8', errors='ignore').strip()
                    if linea:
                        socketio.emit('arduino_msg', {'msg': linea})
                        if linea == 'VISION:STARTED:BTN':
                            estado['activo'] = True
                            registrar_evento('Iniciado por boton fisico', 'success')
                            socketio.emit('cambio_estado', {'estado': 'en_ejecucion'})
                        elif linea == 'VISION:STOPPED:BTN':
                            estado['activo'] = False
                            registrar_evento('Detenido por boton fisico', 'warning')
                            socketio.emit('cambio_estado', {'estado': 'detenido'})
            except Exception:
                pass
        time.sleep(0.01)

# ─── Helpers ──────────────────────────────────────────────────────────────────
def hash_password(pw):
    return hashlib.sha256(pw.encode()).hexdigest()

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'usuario' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

def registrar_evento(mensaje, tipo='info'):
    evento = {'timestamp': datetime.now().isoformat(), 'mensaje': mensaje, 'tipo': tipo}
    eventos.appendleft(evento)
    socketio.emit('nuevo_evento', evento)

def registrar_alarma(mensaje, tipo='warning'):
    global alarma_id_counter
    alarma_id_counter += 1
    alarma = {'id': alarma_id_counter, 'timestamp': datetime.now().isoformat(),
              'mensaje': mensaje, 'tipo': tipo}
    alarmas.append(alarma)
    socketio.emit('nueva_alarma', alarma)



# ─── Hilo de captura + inferencia ─────────────────────────────────────────────
def hilo_inferencia():
    global frame_actual

    # Abrir cámara según plataforma
    if ES_RASPBERRY:
        try:
            from picamera2 import Picamera2
            picam = Picamera2()
            config = picam.create_preview_configuration(
                main={"size": RESOLUCION, "format": "RGB888"}
            )
            picam.configure(config)
            picam.start()
            time.sleep(1)
            print(f'✅ Cámara CSI abierta ({RESOLUCION[0]}x{RESOLUCION[1]})')
            usar_picamera = True
        except Exception as e:
            print(f'❌ Error abriendo cámara CSI: {e}')
            return
    else:
        cap = cv2.VideoCapture(CAMARA_INDEX)
        cap.set(cv2.CAP_PROP_FRAME_WIDTH, RESOLUCION[0])
        cap.set(cv2.CAP_PROP_FRAME_HEIGHT, RESOLUCION[1])
        if not cap.isOpened():
            print('❌ No se pudo abrir la cámara')
            return
        print(f'✅ Cámara abierta ({RESOLUCION[0]}x{RESOLUCION[1]})')
        usar_picamera = False

    fps_counter = 0
    fps_timer = time.time()
    ultimo_emit = time.time()

    while True:
        fps_obj = estado['config']['fps_objetivo']
        intervalo = 1.0 / fps_obj

        t0 = time.time()

        if usar_picamera:
            frame = picam.capture_array()
            frame = cv2.cvtColor(frame, cv2.COLOR_RGB2BGR)
        else:
            ret, frame = cap.read()
            if not ret:
                time.sleep(0.1)
                continue

        display = frame.copy()

        # ─── Inferencia (solo si está activo) ─────────────────────────
        if estado['activo'] and modelo is not None:
            global objetos_previos, siguiente_obj_id
            umbral_bajo = estado['config']['umbral_confianza']
            umbral_alto = estado['config']['umbral_aceptacion']

            results = modelo.predict(frame, conf=umbral_bajo, imgsz=640, verbose=False)

            # Recopilar todas las detecciones del frame actual
            detecciones_frame = []
            for result in results:
                boxes = result.boxes
                if boxes is None or len(boxes) == 0:
                    continue
                for box in boxes:
                    cls_raw = int(box.cls[0])
                    conf = float(box.conf[0])
                    x1, y1, x2, y2 = map(int, box.xyxy[0])
                    if conf >= umbral_alto:
                        cls_id = cls_raw
                    else:
                        cls_id = -1
                    detecciones_frame.append({
                        'bbox': [x1, y1, x2, y2],
                        'cls_id': cls_id,
                        'conf': conf,
                    })

            # ── Tracking por IoU (flanco de subida) ───────────────────
            # Asociar detecciones actuales con objetos previamente trackeados
            usados_prev = set()   # Índices de objetos previos ya asociados
            usados_det  = set()   # Índices de detecciones ya asociadas
            objetos_actuales = []

            # Calcular matriz de IoU entre previos y actuales
            for i_det, det in enumerate(detecciones_frame):
                mejor_iou = 0
                mejor_j = -1
                for j_prev, prev in enumerate(objetos_previos):
                    if j_prev in usados_prev:
                        continue
                    iou = calcular_iou(det['bbox'], prev['bbox'])
                    if iou > mejor_iou:
                        mejor_iou = iou
                        mejor_j = j_prev

                if mejor_iou >= IOU_UMBRAL and mejor_j >= 0:
                    # Mismo objeto que ya existía → NO contar de nuevo
                    usados_prev.add(mejor_j)
                    usados_det.add(i_det)
                    objetos_actuales.append({
                        'id': objetos_previos[mejor_j]['id'],
                        'bbox': det['bbox'],
                        'cls_id': det['cls_id'],
                        'conf': det['conf'],
                        'frames_sin_ver': 0,
                        'es_nuevo': False,
                    })
                else:
                    # Objeto NUEVO → flanco de subida → CONTAR
                    objetos_actuales.append({
                        'id': siguiente_obj_id,
                        'bbox': det['bbox'],
                        'cls_id': det['cls_id'],
                        'conf': det['conf'],
                        'frames_sin_ver': 0,
                        'es_nuevo': True,
                    })
                    siguiente_obj_id += 1

            # Mantener objetos previos no asociados por unos frames más
            for j_prev, prev in enumerate(objetos_previos):
                if j_prev not in usados_prev:
                    prev['frames_sin_ver'] += 1
                    if prev['frames_sin_ver'] < MAX_FRAMES_AUSENTE:
                        prev['es_nuevo'] = False
                        objetos_actuales.append(prev)

            objetos_previos = objetos_actuales

            # ── Dibujar y contar solo objetos NUEVOS ──────────────────
            for obj in objetos_actuales:
                x1, y1, x2, y2 = obj['bbox']
                cls_id = obj['cls_id']
                conf = obj['conf']
                nombre = NOMBRES_CLASE.get(cls_id, '???')
                color = COLORES_BGR.get(cls_id, (255, 255, 255))
                label = f'{nombre} {conf:.0%}'

                # Dibujar bbox (siempre, para todos los objetos visibles)
                cv2.rectangle(display, (x1, y1), (x2, y2), color, 2)
                (tw, th), _ = cv2.getTextSize(label, cv2.FONT_HERSHEY_SIMPLEX, 0.6, 2)
                cv2.rectangle(display, (x1, y1 - th - 10), (x1 + tw + 6, y1), color, -1)
                cv2.putText(display, label, (x1 + 3, y1 - 5),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 0, 0), 2)

                # Solo contar en el flanco de subida (primera aparición)
                if obj['es_nuevo']:
                    estado['contadores']['inspeccionadas'] += 1
                    if cls_id == 0:  # Tornillo
                        estado['contadores']['aceptadas'] += 1
                    elif cls_id == 1:  # Tuerca
                        estado['contadores']['aceptadas'] += 1
                    elif cls_id == -1:  # Objeto extraño
                        estado['contadores']['rechazadas'] += 1
                        estado['contadores']['objetos_extranos'] += 1
                        registrar_alarma(f'Objeto extraño detectado ({conf:.0%})', 'error')

                    registrar_evento(f'{nombre} detectado ({conf:.0%})',
                                     'success' if cls_id >= 0 else 'error')

                    # Mover servo según clase detectada
                    enviar_comando(f'SERVO:{cls_id}')

                    detecciones_log.append({
                        'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                        'clase': nombre,
                        'confianza': round(conf * 100, 1),
                        'resultado': 'Aceptada' if cls_id >= 0 else 'Rechazada',
                        'bbox': f'({x1},{y1})-({x2},{y2})',
                    })

            # Métricas horarias
            hora = datetime.now().strftime('%H:00')
            if hora not in metricas_horarias:
                metricas_horarias[hora] = {'aceptadas': 0, 'rechazadas': 0}
            metricas_horarias[hora]['aceptadas'] = estado['contadores']['aceptadas']
            metricas_horarias[hora]['rechazadas'] = estado['contadores']['rechazadas']

        # ─── Actualizar frame ─────────────────────────────────────────
        with frame_lock:
            frame_actual = display

        # ─── FPS ──────────────────────────────────────────────────────
        fps_counter += 1
        if time.time() - fps_timer >= 1.0:
            estado['fps'] = fps_counter
            fps_counter = 0
            fps_timer = time.time()

        # ─── Emitir stats cada 500ms ──────────────────────────────────
        if time.time() - ultimo_emit >= 0.5:
            socketio.emit('estadisticas', {
                'contadores': estado['contadores'],
                'fps': estado['fps'],
                'estado': 'en_ejecucion' if estado['activo'] else 'detenido',
            })
            socketio.emit('graficos', {
                'metricas_horarias': [{'hora': k, **v} for k, v in metricas_horarias.items()],
            })
            ultimo_emit = time.time()

        # ─── Limitar FPS ──────────────────────────────────────────────
        elapsed = time.time() - t0
        if elapsed < intervalo:
            time.sleep(intervalo - elapsed)

    if usar_picamera:
        picam.stop()
    else:
        cap.release()

# ─── Stream MJPEG ─────────────────────────────────────────────────────────────
def generar_frames():
    while True:
        with frame_lock:
            if frame_actual is None:
                time.sleep(0.05)
                continue
            _, buffer = cv2.imencode('.jpg', frame_actual, [cv2.IMWRITE_JPEG_QUALITY, 80])
            frame_bytes = buffer.tobytes()

        yield (b'--frame\r\n'
               b'Content-Type: image/jpeg\r\n\r\n' + frame_bytes + b'\r\n')
        time.sleep(1.0 / estado['config']['fps_objetivo'])

# ─── Rutas de Autenticación ───────────────────────────────────────────────────
@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'usuario' in session:
        return redirect(url_for('index'))
    if request.method == 'POST':
        usuario = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        if not usuario or not password:
            flash('Por favor ingresa usuario y contraseña.', 'error')
            return render_template('login.html')
        if usuario in USUARIOS and USUARIOS[usuario] == hash_password(password):
            session['usuario'] = usuario
            session.permanent = True
            registrar_evento(f'Login: {usuario}', 'info')
            return redirect(url_for('index'))
        else:
            flash('Usuario o contraseña incorrectos.', 'error')
    return render_template('login.html')

@app.route('/logout')
def logout():
    usuario = session.pop('usuario', None)
    if usuario:
        registrar_evento(f'Logout: {usuario}', 'info')
    return redirect(url_for('login'))

# ─── Rutas principales ────────────────────────────────────────────────────────
@app.route('/')
@login_required
def index():
    return render_template('index.html', usuario=session.get('usuario'))

@app.route('/video_feed')
@login_required
def video_feed():
    return Response(generar_frames(), mimetype='multipart/x-mixed-replace; boundary=frame')

# ─── API: Control ─────────────────────────────────────────────────────────────
@app.route('/api/control/iniciar', methods=['POST'])
@login_required
def api_iniciar():
    global motorActivo
    estado['activo'] = True
    ok = enviar_comando('START')
    if ok:
        motorActivo = True
    registrar_evento('Sistema iniciado', 'success')
    socketio.emit('cambio_estado', {'estado': 'en_ejecucion'})
    return jsonify({'status': 'ok', 'msg': 'Sistema iniciado'})

@app.route('/api/control/detener', methods=['POST'])
@login_required
def api_detener():
    global motorActivo
    estado['activo'] = False
    ok = enviar_comando('STOP')
    if ok:
        motorActivo = False
    registrar_evento('Sistema detenido', 'warning')
    socketio.emit('cambio_estado', {'estado': 'detenido'})
    return jsonify({'status': 'ok', 'msg': 'Sistema detenido'})

# ─── API: Configuración ──────────────────────────────────────────────────────
@app.route('/api/config', methods=['POST'])
@login_required
def api_config():
    data = request.get_json()
    if 'velocidad_banda' in data:
        estado['config']['velocidad_banda'] = int(data['velocidad_banda'])
        enviar_comando(f"VEL:{data['velocidad_banda']}")
    if 'umbral_confianza' in data:
        estado['config']['umbral_confianza'] = float(data['umbral_confianza'])
    if 'umbral_aceptacion' in data:
        estado['config']['umbral_aceptacion'] = float(data['umbral_aceptacion'])
    if 'fps_objetivo' in data:
        estado['config']['fps_objetivo'] = int(data['fps_objetivo'])
    registrar_evento('Configuración actualizada', 'info')
    socketio.emit('configuracion_actualizada', estado['config'])
    return jsonify({'status': 'ok'})

# ─── API: Datos ───────────────────────────────────────────────────────────────
@app.route('/api/estadisticas')
@login_required
def api_estadisticas():
    return jsonify({
        'contadores': estado['contadores'],
        'fps': estado['fps'],
        'estado': 'en_ejecucion' if estado['activo'] else 'detenido',
    })

@app.route('/api/graficos')
@login_required
def api_graficos():
    return jsonify({
        'metricas_horarias': [{'hora': k, **v} for k, v in metricas_horarias.items()],
    })

@app.route('/api/eventos')
@login_required
def api_eventos():
    return jsonify(list(eventos))

@app.route('/api/alarmas')
@login_required
def api_alarmas():
    return jsonify(alarmas)

@app.route('/api/alarmas/limpiar/<int:alarma_id>', methods=['POST'])
@login_required
def api_limpiar_alarma(alarma_id):
    global alarmas
    alarmas = [a for a in alarmas if a['id'] != alarma_id]
    return jsonify({'success': True})

@app.route('/api/contadores/reset', methods=['POST'])
@login_required
def api_reset_contadores():
    global objetos_previos, siguiente_obj_id
    for k in estado['contadores']:
        estado['contadores'][k] = 0
    metricas_horarias.clear()
    detecciones_log.clear()
    objetos_previos = []
    siguiente_obj_id = 0
    registrar_evento('Contadores reseteados', 'info')
    return jsonify({'status': 'ok', 'msg': 'Contadores reseteados'})

@app.route('/api/status')
@login_required
def api_status():
    return jsonify({
        'activo': estado['activo'],
        'motor': motorActivo,
        'fps': estado['fps'],
        'config': estado['config'],
    })

# ─── API: Exportar a Excel ───────────────────────────────────────────────────
@app.route('/api/exportar')
@login_required
def api_exportar():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()

    # ── Hoja 1: Resumen ──────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Resumen'

    header_font = Font(name='Arial', bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='0D6EFD')
    label_font = Font(name='Arial', bold=True, size=11)
    value_font = Font(name='Arial', size=11)
    border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 18

    # Título
    ws.merge_cells('A1:B1')
    ws['A1'] = 'Reporte Vision — Tuercas y Tornillos'
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color='0D6EFD')

    ws['A3'] = 'Sesión'
    ws['A3'].font = label_font
    ws['B3'] = sesion_inicio.strftime('%Y-%m-%d %H:%M')
    ws['B3'].font = value_font
    ws['A4'] = 'Exportado'
    ws['A4'].font = label_font
    ws['B4'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ws['B4'].font = value_font
    ws['A5'] = 'Usuario'
    ws['A5'].font = label_font
    ws['B5'] = session.get('usuario', '-')
    ws['B5'].font = value_font

    # Contadores
    resumen = [
        ('', ''),
        ('CONTADORES', 'CANTIDAD'),
        ('Piezas Inspeccionadas', estado['contadores']['inspeccionadas']),
        ('Piezas Aceptadas', estado['contadores']['aceptadas']),
        ('Piezas Rechazadas', estado['contadores']['rechazadas']),
        ('Objetos Extraños', estado['contadores']['objetos_extranos']),
    ]

    for i, (lbl, val) in enumerate(resumen, start=7):
        ws.cell(row=i, column=1, value=lbl).font = label_font
        ws.cell(row=i, column=2, value=val).font = value_font
        if lbl == 'CONTADORES':
            ws.cell(row=i, column=1).font = header_font
            ws.cell(row=i, column=1).fill = header_fill
            ws.cell(row=i, column=2).font = header_font
            ws.cell(row=i, column=2).fill = header_fill

    # ── Hoja 2: Detecciones ──────────────────────────────────────────
    ws2 = wb.create_sheet('Detecciones')

    columnas = ['#', 'Fecha/Hora', 'Clase', 'Confianza (%)', 'Resultado', 'Bounding Box']
    anchos = [8, 22, 16, 16, 14, 22]

    for i, (col, ancho) in enumerate(zip(columnas, anchos), start=1):
        cell = ws2.cell(row=1, column=i, value=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
        ws2.column_dimensions[chr(64 + i)].width = ancho

    ok_fill = PatternFill('solid', fgColor='E8F5E9')
    nok_fill = PatternFill('solid', fgColor='FFEBEE')

    for idx, det in enumerate(detecciones_log, start=1):
        row = idx + 1
        fill = ok_fill if det['resultado'] == 'Aceptada' else nok_fill
        datos = [idx, det['timestamp'], det['clase'], det['confianza'],
                 det['resultado'], det['bbox']]
        for col, val in enumerate(datos, start=1):
            cell = ws2.cell(row=row, column=col, value=val)
            cell.font = value_font
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

    # ── Hoja 3: Eventos ──────────────────────────────────────────────
    ws3 = wb.create_sheet('Eventos')

    cols_ev = ['#', 'Fecha/Hora', 'Tipo', 'Mensaje']
    anchos_ev = [8, 22, 12, 50]

    for i, (col, ancho) in enumerate(zip(cols_ev, anchos_ev), start=1):
        cell = ws3.cell(row=1, column=i, value=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
        ws3.column_dimensions[chr(64 + i)].width = ancho

    for idx, ev in enumerate(list(eventos), start=1):
        row = idx + 1
        ts = ev.get('timestamp', '')
        datos = [idx, ts, ev.get('tipo', ''), ev.get('mensaje', '')]
        for col, val in enumerate(datos, start=1):
            cell = ws3.cell(row=row, column=col, value=val)
            cell.font = value_font
            cell.border = border

    # ── Guardar y enviar ─────────────────────────────────────────────
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"vision_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    registrar_evento(f'Reporte exportado: {filename}', 'info')

    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ─── Socket.IO ────────────────────────────────────────────────────────────────
@socketio.on('connect')
def on_connect():
    emit('estadisticas', {
        'contadores': estado['contadores'],
        'fps': estado['fps'],
        'estado': 'en_ejecucion' if estado['activo'] else 'detenido',
    })

@socketio.on('velocidad')
def on_velocidad(data):
    pct = data.get('valor', 50)
    estado['config']['velocidad_banda'] = pct
    enviar_comando(f'VEL:{pct}')

# ─── Main ─────────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    cargar_modelo()
    conectar_arduino()

    t_serial = threading.Thread(target=leer_serial, daemon=True)
    t_serial.start()

    t_video = threading.Thread(target=hilo_inferencia, daemon=True)
    t_video.start()

    print('\n' + '=' * 60)
    print('  Sistema Vision — Servidor listo')
    print(f'  http://localhost:5000')
    print('=' * 60)

    socketio.run(app, host='0.0.0.0', port=5000, debug=False, allow_unsafe_werkzeug=True)



